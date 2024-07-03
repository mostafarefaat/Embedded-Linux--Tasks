
import csv 
import openpyxl
import pandas as pd

            # Simple File Funtions #
# input_List = ['Red', 'Green', 'Blue', 'Violet', 'Balck']

# with open("Python/Session_3/file.txt", "r+") as file:

#     list = file.readlines()
#     print("Number of Lines is: ", end=" ")
#     print(len(list))
#     file.seek(0)
#     String = file.read()
#     print("Number of Words is: ", end=" ")
#     print(len(String.split()))
#     file.write(' '.join(input_List))
#     file.seek(0)
#     print(file.read())

            # Simple CSV Funtions #

# file = open("simple.csv", "r")
# csvFile = csv.reader(file)
# for line in csvFile:
#     print(line)

            # Simple Openpyxl Funtions #
# wb = openpyxl.Workbook() #Create new Workbook
# data = [
#     ['Name', 'Age'],
#     ['Alice', 25],
#     ['Bob', 30],
#     ['Charlie', 35],
# ]

# for row in data:
#     wb.active.append(row)
# wb.save("simple.xlsx")

# wb = openpyxl.load_workbook("simple.xlsx", "a")
# print(wb)
# for row in wb["Sheet1"].iter_rows(values_only=True):
#     print(row)

# wb.close()

            # Simple Pandas Funtions #
# def read_csv(file_path):
#     return pd.read_csv(file_path)

# def read_excel(file_path):
#     return pd.read_excel(file_path)

# data = {
#     'Name' : ['John', 'Jane', 'Mike'],
#     'Age'  : [25, 30 , 35],
#     'City' : ['New York', 'London', 'Paris']
# }

# df = pd.DataFrame(data)
# df.to_csv('Output.csv', index=False)
# df.to_excel('Output.xlsx', index=False)

# print(read_csv('Output.csv'))
# print('-'*30)
# print(read_excel('Output.xlsx'))
